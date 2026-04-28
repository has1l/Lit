"""
Парсинг и индексация документов в ChromaDB.
Поддержка: PDF, DOCX, XLSX/XLS.
"""

import io
import re
import uuid
from pathlib import Path

HEADING_STYLES = {
    "heading 1", "heading 2", "heading 3",
    "заголовок 1", "заголовок 2", "заголовок 3",
}
_HEADING_RE = re.compile(r"^(Статья|Раздел|Глава|Пункт)\s+\d+", re.IGNORECASE)


def _chunk_text(text: str, chunk_size: int = 1000, overlap: int = 200) -> list[str]:
    text = re.sub(r'\s+', ' ', text).strip()
    if not text:
        return []
    chunks, start = [], 0
    while start < len(text):
        end = start + chunk_size
        split = text.rfind('. ', start + overlap, end)
        if split == -1:
            split = text.rfind(' ', start + overlap, end)
        if split == -1:
            split = end
        chunks.append(text[start:split + 1].strip())
        start = split + 1
    return [c for c in chunks if len(c) > 30]


# ── Парсеры ───────────────────────────────────────────────────────────────────

def _parse_pdf_pages(file_bytes: bytes) -> list[dict]:
    """Returns list of {text, page_number}."""
    import pypdf
    reader = pypdf.PdfReader(io.BytesIO(file_bytes))
    return [
        {"text": (page.extract_text() or "").strip(), "page_number": i + 1}
        for i, page in enumerate(reader.pages)
        if (page.extract_text() or "").strip()
    ]


def _parse_docx_sections(file_bytes: bytes) -> list[dict]:
    """Returns list of {text, section, page_number} where page_number=0 (N/A for DOCX)."""
    from docx import Document
    doc = Document(io.BytesIO(file_bytes))

    sections, heading, lines = [], "Общие положения", []

    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue
        style = (para.style.name or "").lower()
        is_h = style in HEADING_STYLES or bool(_HEADING_RE.match(text))
        if is_h:
            if lines:
                sections.append({
                    "text": " ".join(lines),
                    "section": heading,
                    "page_number": 0,
                })
                lines = []
            heading = text[:120]
        else:
            lines.append(text)

    if lines:
        sections.append({"text": " ".join(lines), "section": heading, "page_number": 0})

    return [s for s in sections if len(s["text"]) > 60]


def _parse_xlsx(file_bytes: bytes) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    lines = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            row_text = ' | '.join(str(c) for c in row if c is not None)
            if row_text.strip():
                lines.append(row_text)
    return '\n'.join(lines)


# ── Основная функция ──────────────────────────────────────────────────────────

def ingest_document(
    filename: str,
    file_bytes: bytes,
    uploader_email: str,
    audience: str = 'all',
) -> int:
    """
    Парсит файл, делит на чанки и добавляет в ChromaDB.
    Возвращает количество добавленных чанков.
    """
    ext = Path(filename).suffix.lower()
    doc_title = Path(filename).stem

    # Строим список raw-блоков: {text, section, page_number}
    raw_blocks: list[dict] = []

    if ext == '.pdf':
        for page in _parse_pdf_pages(file_bytes):
            raw_blocks.append({
                "text":        page["text"],
                "section":     f"Страница {page['page_number']}",
                "page_number": page["page_number"],
            })
    elif ext in ('.docx', '.doc'):
        for sec in _parse_docx_sections(file_bytes):
            raw_blocks.append({
                "text":        sec["text"],
                "section":     sec["section"],
                "page_number": 0,
            })
    elif ext in ('.xlsx', '.xls'):
        text = _parse_xlsx(file_bytes)
        if text.strip():
            raw_blocks.append({"text": text, "section": doc_title, "page_number": 0})
    else:
        raise ValueError(f"Неподдерживаемый формат: {ext}")

    if not raw_blocks:
        raise ValueError("Документ пустой или не удалось извлечь текст")

    # Разбиваем каждый блок на чанки
    all_chunks: list[dict] = []
    for block in raw_blocks:
        for chunk_text in _chunk_text(block["text"]):
            all_chunks.append({
                "text":        chunk_text,
                "section":     block["section"],
                "page_number": block["page_number"],
            })

    if not all_chunks:
        raise ValueError("Не удалось разбить текст на чанки")

    # ChromaDB
    chroma_path = Path(__file__).parent / "chroma_db"
    chroma_path.mkdir(exist_ok=True)

    import chromadb
    from sentence_transformers import SentenceTransformer

    client = chromadb.PersistentClient(path=str(chroma_path))
    try:
        coll = client.get_collection("hr_documents")
    except Exception:
        coll = client.create_collection("hr_documents")

    model = SentenceTransformer("paraphrase-multilingual-MiniLM-L12-v2")

    ids, docs, embeddings, metas = [], [], [], []
    for i, chunk in enumerate(all_chunks):
        chunk_id = f"{uuid.uuid4().hex}_{i}"
        emb = model.encode([chunk["text"]])[0].tolist()
        ids.append(chunk_id)
        docs.append(chunk["text"])
        embeddings.append(emb)
        metas.append({
            "title":        doc_title,
            "section":      chunk["section"],
            "audience":     audience,
            "source_file":  filename,
            "chunk_index":  i,
            "uploader":     uploader_email,
            "page_number":  chunk["page_number"],
        })

    coll.add(ids=ids, documents=docs, embeddings=embeddings, metadatas=metas)
    return len(all_chunks)
